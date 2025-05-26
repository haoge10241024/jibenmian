import streamlit as st
import akshare as ak
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from scipy import stats
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import warnings
import io
import base64
from typing import Dict, List, Tuple, Optional
import time
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
import tempfile
import os

warnings.filterwarnings('ignore')

# 设置页面配置
st.set_page_config(
    page_title="期货库存分析系统",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 优化缓存配置
@st.cache_data(ttl=7200, max_entries=1000)  # 缓存2小时，最多1000个条目
def cached_futures_inventory_em(symbol):
    """缓存的期货库存数据获取"""
    return ak.futures_inventory_em(symbol=symbol)

@st.cache_data(ttl=7200, max_entries=1000)  # 缓存2小时，最多1000个条目  
def cached_futures_hist_em(symbol, period="daily"):
    """缓存的期货历史行情数据获取"""
    return ak.futures_hist_em(symbol=symbol, period=period)

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

class FuturesInventoryAnalyzer:
    def __init__(self, confidence_level: float = 0.95):
        self.confidence_level = confidence_level
        self.seasonal_periods = {
            '农产品': 12,
            '工业品': 4,
            '能源化工': 4
        }
        
    def calculate_seasonal_factor(self, df: pd.DataFrame, category: str) -> pd.Series:
        """计算季节性因子"""
        period = self.seasonal_periods.get(category, 12)
        seasonal = df['库存'].rolling(window=period).mean()
        return seasonal
    
    def calculate_inventory_velocity(self, df: pd.DataFrame, days: int = 30) -> float:
        """计算库存周转率"""
        recent_data = df.tail(days)
        return recent_data['增减'].abs().sum() / recent_data['库存'].mean()
    
    def calculate_trend_strength(self, df: pd.DataFrame, window: int = 30) -> float:
        """计算趋势强度"""
        try:
            if len(df) < window:
                return 0.0
            
            price_change = df['库存'].diff().dropna()
            if len(price_change) < window:
                return 0.0
            
            positive_moves = price_change[price_change > 0].rolling(window=window).sum()
            negative_moves = price_change[price_change < 0].rolling(window=window).sum()
            
            if positive_moves.empty or negative_moves.empty:
                return 0.0
            
            last_positive = positive_moves.iloc[-1] if not pd.isna(positive_moves.iloc[-1]) else 0
            last_negative = negative_moves.iloc[-1] if not pd.isna(negative_moves.iloc[-1]) else 0
            
            total_moves = last_positive + abs(last_negative)
            if total_moves == 0:
                return 0.0
            
            return abs(last_positive - abs(last_negative)) / total_moves
        
        except Exception:
            return 0.0
    
    def calculate_dynamic_threshold(self, df: pd.DataFrame, window: int = 60) -> float:
        """计算动态阈值"""
        try:
            volatility = df['增减'].rolling(window=window).std().iloc[-1]
            return volatility * stats.norm.ppf(self.confidence_level)
        except:
            return 0.0
    
    def analyze_inventory_trend(self, df: pd.DataFrame, category: str) -> Dict:
        """综合分析库存趋势"""
        try:
            recent_data = df.tail(30)
            total_change = recent_data['增减'].sum()
            avg_change = total_change / len(recent_data)
            
            # 优化变化率计算
            start_inventory = recent_data['库存'].iloc[0]
            end_inventory = recent_data['库存'].iloc[-1]
            
            # 更稳健的变化率计算
            if start_inventory > 0:
                change_rate = (end_inventory - start_inventory) / start_inventory * 100
                # 对于极端值进行合理限制
                if abs(change_rate) > 200:  # 如果变化率超过200%，可能是数据异常
                    # 使用平均库存作为基准重新计算
                    avg_inventory = recent_data['库存'].mean()
                    if avg_inventory > 0:
                        change_rate = (end_inventory - start_inventory) / avg_inventory * 100
                    else:
                        change_rate = 0
            else:
                # 起始库存为0时，使用不同的计算方法
                if end_inventory > 0:
                    change_rate = 100  # 从0增加到有库存，设为100%
                else:
                    change_rate = 0
            
            # 限制变化率范围到合理区间
            change_rate = min(max(change_rate, -150), 150)  # 限制在-150%到150%之间
            
            seasonal_factor = self.calculate_seasonal_factor(df, category)
            inventory_velocity = self.calculate_inventory_velocity(df)
            trend_strength = self.calculate_trend_strength(df)
            dynamic_threshold = self.calculate_dynamic_threshold(df)
            
            # 趋势判断 - 优化判断逻辑
            trend = '稳定'
            # 主要基于变化率和平均日变化，趋势强度作为辅助判断
            if abs(change_rate) > 15:  # 变化率超过15%
                if change_rate > 15 and avg_change > 0:
                    trend = '累库'
                elif change_rate < -15 and avg_change < 0:
                    trend = '去库'
            elif abs(change_rate) > 8:  # 变化率在8-15%之间，需要趋势强度支持
                if change_rate > 8 and avg_change > 0 and trend_strength > 0.1:
                    trend = '累库'
                elif change_rate < -8 and avg_change < 0 and trend_strength > 0.1:
                    trend = '去库'
            
            signal_strength = min(abs(change_rate) / max(dynamic_threshold, 1), 1.0)
            
            return {
                '趋势': trend,
                '变化率': change_rate,
                '平均日变化': avg_change,
                '趋势强度': trend_strength,
                '信号强度': signal_strength,
                '库存周转率': inventory_velocity,
                '季节性因子': seasonal_factor.iloc[-1] if not seasonal_factor.empty else 0,
                '动态阈值': dynamic_threshold
            }
        except Exception:
            return {
                '趋势': '稳定',
                '变化率': 0,
                '平均日变化': 0,
                '趋势强度': 0,
                '信号强度': 0,
                '库存周转率': 0,
                '季节性因子': 0,
                '动态阈值': 0
            }

def get_futures_category(symbol: str) -> str:
    """获取期货品种分类"""
    categories = {
        '农产品': ['豆一', '豆二', '豆粕', '豆油', '玉米', '玉米淀粉', '菜粕', '菜油', '棕榈', '白糖', '棉花', '苹果'],
        '工业品': ['螺纹钢', '热卷', '铁矿石', '焦煤', '焦炭', '不锈钢', '沪铜', '沪铝', '沪锌', '沪铅', '沪镍', '沪锡'],
        '能源化工': ['原油', '燃油', '沥青', 'PTA', '甲醇', '乙二醇', 'PVC', 'PP', '塑料', '橡胶', '20号胶']
    }
    
    for category, symbols in categories.items():
        if symbol in symbols:
            return category
    return '其他'

import concurrent.futures

def get_single_inventory_data_streamlit(symbol: str) -> Optional[pd.DataFrame]:
    """获取单个期货品种的库存数据（Streamlit版本）"""
    try:
        df = cached_futures_inventory_em(symbol)
        
        if df is not None and not df.empty and '日期' in df.columns and '库存' in df.columns:
            df['日期'] = pd.to_datetime(df['日期'])
            df['库存'] = pd.to_numeric(df['库存'], errors='coerce')
            df = df.dropna(subset=['日期', '库存'])
            
            if len(df) >= 2:
                df['增减'] = df['库存'].diff()
                df = df.dropna(subset=['增减'])
                
                if len(df) >= 30:
                    return df
        return None
    except Exception:
        return None

@st.cache_data(ttl=3600)  # 缓存1小时
def get_futures_inventory_data(symbols_list):
    """并行获取期货库存数据"""
    data_dict = {}
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    status_text.text(f"开始并行获取 {len(symbols_list)} 个品种的库存数据...")
    
    # 使用线程池并行获取数据
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        # 提交所有任务
        future_to_symbol = {
            executor.submit(get_single_inventory_data_streamlit, symbol): symbol 
            for symbol in symbols_list
        }
        
        # 收集结果
        completed = 0
        for future in concurrent.futures.as_completed(future_to_symbol):
            symbol = future_to_symbol[future]
            completed += 1
            
            try:
                df = future.result()
                if df is not None:
                    data_dict[symbol] = df
                    status_text.text(f"✓ {symbol} 数据获取成功 ({completed}/{len(symbols_list)})")
                else:
                    status_text.text(f"✗ {symbol} 数据获取失败 ({completed}/{len(symbols_list)})")
            except Exception:
                status_text.text(f"✗ {symbol} 处理异常 ({completed}/{len(symbols_list)})")
            
            progress_bar.progress(completed / len(symbols_list))
            time.sleep(0.05)  # 短暂延迟以显示进度
    
    progress_bar.empty()
    status_text.empty()
    return data_dict

@st.cache_data(ttl=3600)  # 缓存1小时
def get_futures_price_data_cached(symbol: str) -> Optional[pd.DataFrame]:
    """
    获取期货价格数据（带缓存）
    """
    try:
        # 将库存数据的symbol转换为价格数据的symbol（添加"主连"）
        price_symbol = f"{symbol}主连"
        
        # 获取期货历史行情数据
        price_df = cached_futures_hist_em(price_symbol, "daily")
        
        if price_df is None or price_df.empty:
            return None
            
        # 数据预处理
        price_df['时间'] = pd.to_datetime(price_df['时间'])
        price_df = price_df.rename(columns={'时间': '日期', '收盘': '价格'})
        
        # 选择需要的列
        price_df = price_df[['日期', '价格', '开盘', '最高', '最低', '成交量', '持仓量']].copy()
        price_df['价格'] = pd.to_numeric(price_df['价格'], errors='coerce')
        price_df = price_df.dropna(subset=['日期', '价格'])
        
        return price_df
        
    except Exception:
        return None

def get_futures_price_data(symbol: str) -> Optional[pd.DataFrame]:
    """
    获取期货价格数据
    """
    return get_futures_price_data_cached(symbol)

@st.cache_data(ttl=3600)  # 缓存1小时
def get_multiple_price_data_streamlit(symbols_tuple):
    """
    并行获取多个品种的价格数据（Streamlit版本）
    注意：使用tuple作为参数以支持缓存
    """
    symbols = list(symbols_tuple)
    price_data_dict = {}
    
    if not symbols:
        return price_data_dict
    
    # 创建进度显示
    progress_bar = st.progress(0)
    status_text = st.empty()
    status_text.text(f"开始并行获取 {len(symbols)} 个品种的价格数据...")
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=6) as executor:
        # 提交所有任务
        future_to_symbol = {
            executor.submit(get_futures_price_data_cached, symbol): symbol 
            for symbol in symbols
        }
        
        # 收集结果
        completed = 0
        for future in concurrent.futures.as_completed(future_to_symbol):
            symbol = future_to_symbol[future]
            completed += 1
            
            try:
                price_df = future.result()
                if price_df is not None:
                    price_data_dict[symbol] = price_df
                    status_text.text(f"✓ {symbol}主连 价格数据获取成功 ({completed}/{len(symbols)})")
                else:
                    status_text.text(f"✗ {symbol}主连 价格数据获取失败 ({completed}/{len(symbols)})")
            except Exception:
                status_text.text(f"✗ {symbol}主连 处理异常 ({completed}/{len(symbols)})")
            
            progress_bar.progress(completed / len(symbols))
            time.sleep(0.05)
    
    progress_bar.empty()
    status_text.empty()
    return price_data_dict

def align_inventory_and_price_data(inventory_df: pd.DataFrame, price_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    对齐库存数据和价格数据的时间范围
    """
    try:
        # 确保日期列是datetime类型
        inventory_df['日期'] = pd.to_datetime(inventory_df['日期'])
        price_df['日期'] = pd.to_datetime(price_df['日期'])
        
        # 找到两个数据集的共同时间范围
        inventory_start = inventory_df['日期'].min()
        inventory_end = inventory_df['日期'].max()
        price_start = price_df['日期'].min()
        price_end = price_df['日期'].max()
        
        # 取交集时间范围
        common_start = max(inventory_start, price_start)
        common_end = min(inventory_end, price_end)
        
        # 过滤数据到共同时间范围
        aligned_inventory = inventory_df[
            (inventory_df['日期'] >= common_start) & 
            (inventory_df['日期'] <= common_end)
        ].copy()
        
        aligned_price = price_df[
            (price_df['日期'] >= common_start) & 
            (price_df['日期'] <= common_end)
        ].copy()
        
        return aligned_inventory, aligned_price
        
    except Exception as e:
        st.warning(f"数据对齐失败: {str(e)}")
        return inventory_df, price_df

def create_plotly_trend_chart(df, symbol, analysis_result):
    """创建Plotly交互式趋势图"""
    fig = make_subplots(
        rows=2, cols=1,
        subplot_titles=(f'{symbol} 库存趋势', '库存增减'),
        vertical_spacing=0.1,
        row_heights=[0.7, 0.3]
    )
    
    # 库存趋势线
    fig.add_trace(
        go.Scatter(
            x=df['日期'],
            y=df['库存'],
            mode='lines',
            name='库存',
            line=dict(color='blue', width=2),
            yaxis='y'
        ),
        row=1, col=1
    )
    
    # 30日移动平均
    ma30 = df['库存'].rolling(window=30).mean()
    fig.add_trace(
        go.Scatter(
            x=df['日期'],
            y=ma30,
            mode='lines',
            name='库存30日均线',
            line=dict(color='red', dash='dash'),
            yaxis='y'
        ),
        row=1, col=1
    )
    
    # 增减柱状图
    colors = ['green' if x > 0 else 'red' for x in df['增减']]
    fig.add_trace(
        go.Bar(
            x=df['日期'],
            y=df['增减'],
            name='日增减',
            marker_color=colors,
            opacity=0.7
        ),
        row=2, col=1
    )
    
    # 更新布局
    fig.update_layout(
        height=600,
        title=f"{symbol} 库存分析 - 趋势: {analysis_result['趋势']} | 变化率: {analysis_result['变化率']:.1f}%",
        showlegend=True
    )
    
    fig.update_xaxes(title_text="日期", row=2, col=1)
    fig.update_yaxes(title_text="库存量", row=1, col=1)
    fig.update_yaxes(title_text="增减量", row=2, col=1)
    
    return fig

def create_plotly_inventory_price_chart(inventory_df, price_df, symbol, analysis_result):
    """创建库存与价格对比的交互式图表"""
    fig = make_subplots(
        rows=3, cols=1,
        subplot_titles=(f'{symbol} 库存与价格走势对比', '库存增减', '价格变化'),
        vertical_spacing=0.08,
        row_heights=[0.5, 0.25, 0.25],
        specs=[[{"secondary_y": True}], [{"secondary_y": False}], [{"secondary_y": False}]]
    )
    
    # 库存趋势线（左y轴）
    fig.add_trace(
        go.Scatter(
            x=inventory_df['日期'],
            y=inventory_df['库存'],
            mode='lines',
            name='库存',
            line=dict(color='blue', width=2.5),
            yaxis='y'
        ),
        row=1, col=1, secondary_y=False
    )
    
    # 库存30日移动平均线
    ma30_inventory = inventory_df['库存'].rolling(window=30).mean()
    fig.add_trace(
        go.Scatter(
            x=inventory_df['日期'],
            y=ma30_inventory,
            mode='lines',
            name='库存30日均线',
            line=dict(color='lightblue', dash='dash', width=1.5),
            yaxis='y'
        ),
        row=1, col=1, secondary_y=False
    )
    
    # 价格趋势线（右y轴）
    if price_df is not None and len(price_df) > 0:
        fig.add_trace(
            go.Scatter(
                x=price_df['日期'],
                y=price_df['价格'],
                mode='lines',
                name='价格',
                line=dict(color='purple', width=2.5),
                yaxis='y2'
            ),
            row=1, col=1, secondary_y=True
        )
        
        # 价格30日移动平均线
        ma30_price = price_df['价格'].rolling(window=30).mean()
        fig.add_trace(
            go.Scatter(
                x=price_df['日期'],
                y=ma30_price,
                mode='lines',
                name='价格30日均线',
                line=dict(color='plum', dash='dash', width=1.5),
                yaxis='y2'
            ),
            row=1, col=1, secondary_y=True
        )
        
        # 计算价格变化率
        if len(price_df) > 1:
            price_change_rate = (price_df['价格'].iloc[-1] - price_df['价格'].iloc[0]) / price_df['价格'].iloc[0] * 100
        else:
            price_change_rate = 0
    else:
        price_change_rate = 0
    
    # 库存增减柱状图
    colors = ['green' if x > 0 else 'red' for x in inventory_df['增减']]
    fig.add_trace(
        go.Bar(
            x=inventory_df['日期'],
            y=inventory_df['增减'],
            name='库存日增减',
            marker_color=colors,
            opacity=0.7
        ),
        row=2, col=1
    )
    
    # 价格变化柱状图
    if price_df is not None and len(price_df) > 1:
        price_change = price_df['价格'].diff()
        price_colors = ['green' if x > 0 else 'red' for x in price_change]
        fig.add_trace(
            go.Bar(
                x=price_df['日期'],
                y=price_change,
                name='价格日变化',
                marker_color=price_colors,
                opacity=0.7
            ),
            row=3, col=1
        )
    
    # 设置y轴标签
    fig.update_yaxes(title_text="库存量", row=1, col=1, secondary_y=False)
    fig.update_yaxes(title_text="价格", row=1, col=1, secondary_y=True)
    fig.update_yaxes(title_text="库存增减", row=2, col=1)
    fig.update_yaxes(title_text="价格变化", row=3, col=1)
    
    # 设置x轴标签
    fig.update_xaxes(title_text="日期", row=3, col=1)
    
    # 更新布局
    title_text = f"{symbol} 库存价格分析 - 趋势: {analysis_result['趋势']} | 库存变化: {analysis_result['变化率']:.1f}%"
    if price_df is not None:
        title_text += f" | 价格变化: {price_change_rate:.1f}%"
    
    fig.update_layout(
        height=800,
        title=title_text,
        showlegend=True,
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1
        )
    )
    
    return fig

def create_summary_charts(results_df, inventory_trends):
    """创建汇总图表"""
    col1, col2 = st.columns(2)
    
    with col1:
        # 趋势分布饼图
        trend_counts = [
            len(inventory_trends['累库品种']),
            len(inventory_trends['去库品种']),
            len(inventory_trends['库存稳定品种'])
        ]
        labels = ['累库品种', '去库品种', '稳定品种']
        colors = ['green', 'red', 'gray']
        
        fig_pie = go.Figure(data=[go.Pie(
            labels=labels,
            values=trend_counts,
            marker_colors=colors,
            textinfo='label+percent'
        )])
        fig_pie.update_layout(title="库存趋势分布")
        st.plotly_chart(fig_pie, use_container_width=True, key=f"summary_pie_chart_{hash(str(len(results_df))) % 1000}")
    
    with col2:
        # 信号强度分布
        fig_hist = px.histogram(
            results_df,
            x='信号强度',
            nbins=20,
            title='信号强度分布',
            color_discrete_sequence=['skyblue']
        )
        st.plotly_chart(fig_hist, use_container_width=True, key=f"summary_hist_chart_{hash(str(len(results_df))) % 1000}")

def create_excel_report(results_df, inventory_trends, data_dict):
    """创建Excel格式的分析报告"""
    # 创建工作簿
    wb = Workbook()
    
    # 删除默认工作表
    wb.remove(wb.active)
    
    # 1. 创建汇总分析工作表
    ws_summary = wb.create_sheet("汇总分析")
    
    # 设置标题
    ws_summary['A1'] = "期货库存分析报告"
    ws_summary['A1'].font = Font(size=16, bold=True)
    ws_summary['A1'].alignment = Alignment(horizontal='center')
    ws_summary.merge_cells('A1:H1')
    
    # 分析时间
    ws_summary['A3'] = f"分析时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary['A3'].font = Font(bold=True)
    
    # 汇总统计
    ws_summary['A5'] = "汇总统计"
    ws_summary['A5'].font = Font(size=14, bold=True)
    
    summary_data = [
        ["指标", "数量", "占比"],
        ["总品种数", len(results_df), "100%"],
        ["累库品种", len(inventory_trends['累库品种']), f"{len(inventory_trends['累库品种'])/len(results_df)*100:.1f}%"],
        ["去库品种", len(inventory_trends['去库品种']), f"{len(inventory_trends['去库品种'])/len(results_df)*100:.1f}%"],
        ["稳定品种", len(inventory_trends['库存稳定品种']), f"{len(inventory_trends['库存稳定品种'])/len(results_df)*100:.1f}%"]
    ]
    
    for row_idx, row_data in enumerate(summary_data, start=6):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 6:  # 标题行
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 2. 创建详细分析结果工作表
    ws_details = wb.create_sheet("详细分析结果")
    
    # 写入列标题
    headers = list(results_df.columns)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_details.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # 写入数据
    for row_idx, (_, row) in enumerate(results_df.iterrows(), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws_details.cell(row=row_idx, column=col_idx, value=value)
    
    # 3. 创建重点品种工作表
    if inventory_trends['累库品种'] or inventory_trends['去库品种']:
        ws_focus = wb.create_sheet("重点关注品种")
        
        current_row = 1
        
        # 累库品种
        if inventory_trends['累库品种']:
            ws_focus.cell(row=current_row, column=1, value="累库品种 TOP10").font = Font(size=14, bold=True)
            current_row += 1
            
            # 标题行
            headers = ["品种", "变化率(%)", "信号强度", "趋势强度", "分类"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_focus.cell(row=current_row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            current_row += 1
            
            # 数据行
            top_accumulation = results_df[results_df['趋势'] == '累库'].head(10)
            for _, row in top_accumulation.iterrows():
                ws_focus.cell(row=current_row, column=1, value=row['品种'])
                ws_focus.cell(row=current_row, column=2, value=f"{row['变化率']:.2f}")
                ws_focus.cell(row=current_row, column=3, value=f"{row['信号强度']:.3f}")
                ws_focus.cell(row=current_row, column=4, value=f"{row['趋势强度']:.3f}")
                ws_focus.cell(row=current_row, column=5, value=row['分类'])
                current_row += 1
            
            current_row += 2
        
        # 去库品种
        if inventory_trends['去库品种']:
            ws_focus.cell(row=current_row, column=1, value="去库品种 TOP10").font = Font(size=14, bold=True)
            current_row += 1
            
            # 标题行
            headers = ["品种", "变化率(%)", "信号强度", "趋势强度", "分类"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_focus.cell(row=current_row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            current_row += 1
            
            # 数据行
            top_depletion = results_df[results_df['趋势'] == '去库'].head(10)
            for _, row in top_depletion.iterrows():
                ws_focus.cell(row=current_row, column=1, value=row['品种'])
                ws_focus.cell(row=current_row, column=2, value=f"{row['变化率']:.2f}")
                ws_focus.cell(row=current_row, column=3, value=f"{row['信号强度']:.3f}")
                ws_focus.cell(row=current_row, column=4, value=f"{row['趋势强度']:.3f}")
                ws_focus.cell(row=current_row, column=5, value=row['分类'])
                current_row += 1
    
    # 4. 创建原始数据工作表（选择性添加）
    if len(data_dict) <= 10:  # 只有在品种数量不太多时才添加原始数据
        for symbol, df in list(data_dict.items())[:5]:  # 最多添加5个品种的原始数据
            ws_data = wb.create_sheet(f"{symbol}_原始数据")
            
            # 写入列标题
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_data.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 写入数据（最近100条记录）
            recent_data = df.tail(100)
            for row_idx, (_, row) in enumerate(recent_data.iterrows(), start=2):
                for col_idx, value in enumerate(row, start=1):
                    if isinstance(value, pd.Timestamp):
                        value = value.strftime('%Y-%m-%d')
                    ws_data.cell(row=row_idx, column=col_idx, value=value)
    
    # 调整列宽
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # 跳过合并单元格
                    if hasattr(cell, 'column_letter'):
                        column_letter = cell.column_letter
                    elif hasattr(cell, 'column'):
                        from openpyxl.utils import get_column_letter
                        column_letter = get_column_letter(cell.column)
                    
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

def create_charts_zip(results_df, inventory_trends, data_dict):
    """创建包含所有图表的ZIP文件"""
    try:
        # 创建临时目录
        with tempfile.TemporaryDirectory() as temp_dir:
            chart_files = []
            
            # 1. 创建汇总图表
            try:
                # 趋势分布饼图
                trend_counts = [
                    len(inventory_trends['累库品种']),
                    len(inventory_trends['去库品种']),
                    len(inventory_trends['库存稳定品种'])
                ]
                labels = ['累库品种', '去库品种', '稳定品种']
                colors = ['green', 'red', 'gray']
                
                fig_pie = go.Figure(data=[go.Pie(
                    labels=labels,
                    values=trend_counts,
                    marker_colors=colors,
                    textinfo='label+percent'
                )])
                fig_pie.update_layout(title="库存趋势分布", width=800, height=600)
                
                pie_path = os.path.join(temp_dir, "01_库存趋势分布.png")
                fig_pie.write_image(pie_path)
                chart_files.append(pie_path)
            except Exception as e:
                print(f"创建趋势分布图失败: {e}")
            
            try:
                # 信号强度分布直方图
                fig_hist = px.histogram(
                    results_df,
                    x='信号强度',
                    nbins=20,
                    title='信号强度分布',
                    color_discrete_sequence=['skyblue']
                )
                fig_hist.update_layout(width=800, height=600)
                
                hist_path = os.path.join(temp_dir, "02_信号强度分布.png")
                fig_hist.write_image(hist_path)
                chart_files.append(hist_path)
            except Exception as e:
                print(f"创建信号强度分布图失败: {e}")
            
            # 2. 创建重点品种图表
            chart_count = 3
            
            # 累库品种图表
            if inventory_trends['累库品种']:
                top_accumulation = results_df[results_df['趋势'] == '累库'].head(5)
                for _, row in top_accumulation.iterrows():
                    try:
                        symbol = row['品种']
                        if symbol in data_dict:
                            df = data_dict[symbol]
                            analysis_result = row.to_dict()
                            
                            # 创建库存趋势图
                            fig = create_plotly_trend_chart(df, symbol, analysis_result)
                            fig.update_layout(width=1000, height=700)
                            
                            chart_path = os.path.join(temp_dir, f"{chart_count:02d}_{symbol}_累库趋势图.png")
                            fig.write_image(chart_path)
                            chart_files.append(chart_path)
                            chart_count += 1
                    except Exception as e:
                        print(f"创建{symbol}累库趋势图失败: {e}")
            
            # 去库品种图表
            if inventory_trends['去库品种']:
                top_depletion = results_df[results_df['趋势'] == '去库'].head(5)
                for _, row in top_depletion.iterrows():
                    try:
                        symbol = row['品种']
                        if symbol in data_dict:
                            df = data_dict[symbol]
                            analysis_result = row.to_dict()
                            
                            # 创建库存趋势图
                            fig = create_plotly_trend_chart(df, symbol, analysis_result)
                            fig.update_layout(width=1000, height=700)
                            
                            chart_path = os.path.join(temp_dir, f"{chart_count:02d}_{symbol}_去库趋势图.png")
                            fig.write_image(chart_path)
                            chart_files.append(chart_path)
                            chart_count += 1
                    except Exception as e:
                        print(f"创建{symbol}去库趋势图失败: {e}")
            
            # 创建ZIP文件
            if chart_files:  # 只有在有图表文件时才创建ZIP
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    for chart_file in chart_files:
                        if os.path.exists(chart_file):
                            zip_file.write(chart_file, os.path.basename(chart_file))
                
                zip_buffer.seek(0)
                return zip_buffer.getvalue()
            else:
                # 如果没有图表文件，创建一个包含说明的ZIP
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    zip_file.writestr("说明.txt", "图表生成失败，请检查数据和环境配置。")
                zip_buffer.seek(0)
                return zip_buffer.getvalue()
                
    except Exception as e:
        print(f"创建图表ZIP包失败: {e}")
        # 返回一个包含错误信息的ZIP文件
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            zip_file.writestr("错误信息.txt", f"图表包生成失败: {str(e)}")
        zip_buffer.seek(0)
        return zip_buffer.getvalue()

def main():
    st.title("📊 期货库存分析系统")
    st.markdown("---")
    
    # 显示使用说明
    if not st.session_state.get('analysis_completed', False):
        st.info("👈 请在左侧配置分析参数，然后点击'开始分析'按钮开始分析。分析完成后，您可以在信号品种对比分析中自由选择品种，页面不会重新加载。")
    
    # 初始化session state
    if 'analysis_completed' not in st.session_state:
        st.session_state.analysis_completed = False
    if 'results_df' not in st.session_state:
        st.session_state.results_df = None
    if 'data_dict' not in st.session_state:
        st.session_state.data_dict = None
    if 'inventory_trends' not in st.session_state:
        st.session_state.inventory_trends = None
    if 'analysis_mode' not in st.session_state:
        st.session_state.analysis_mode = "全品种分析"
    
    # 侧边栏配置
    st.sidebar.header("分析配置")
    
    # 期货品种选择
    all_symbols = [
        "沪铜", "镍", "锡", "沪铝", "苯乙烯", "液化石油气", "低硫燃料油", "棉纱",
        "不锈钢", "短纤", "沪铅", "多晶硅", "丁二烯橡胶", "沪锌", "硅铁", "鸡蛋",
        "瓶片", "工业硅", "沥青", "20号胶", "原木", "豆一", "玉米", "燃油",
        "菜籽", "碳酸锂", "纸浆", "玉米淀粉", "沪银", "沪金", "塑料", "聚丙烯",
        "铁矿石", "豆二", "豆粕", "棕榈", "玻璃", "豆油", "橡胶", "烧碱",
        "菜粕", "PTA", "纯碱", "对二甲苯", "菜油", "生猪", "尿素", "PVC",
        "乙二醇", "氧化铝", "焦炭", "郑棉", "甲醇", "白糖", "锰硅", "焦煤",
        "红枣", "螺纹钢", "花生", "苹果", "热卷"
    ]
    
    analysis_mode = st.sidebar.radio(
        "选择分析模式",
        ["全品种分析", "自定义品种分析", "单品种详细分析"],
        key="analysis_mode_radio"
    )
    
    if analysis_mode == "全品种分析":
        selected_symbols = all_symbols
    elif analysis_mode == "自定义品种分析":
        selected_symbols = st.sidebar.multiselect(
            "选择要分析的品种",
            all_symbols,
            default=all_symbols[:10],
            key="custom_symbols_select"
        )
    else:  # 单品种详细分析
        selected_symbols = [st.sidebar.selectbox("选择品种", all_symbols, key="single_symbol_select")]
    
    # 分析参数
    st.sidebar.subheader("分析参数")
    confidence_level = st.sidebar.slider("置信水平", 0.90, 0.99, 0.95, 0.01, key="confidence_slider")
    change_threshold = st.sidebar.slider("变化率阈值 (%)", 5, 20, 10, 1, key="change_threshold_slider")
    trend_threshold = st.sidebar.slider("趋势强度阈值", 0.1, 0.8, 0.3, 0.1, key="trend_threshold_slider")
    
    # 重置分析按钮
    if st.sidebar.button("🔄 重新分析", type="secondary"):
        # 清理所有分析相关的session state
        keys_to_clear = [key for key in st.session_state.keys() if 
                        key.startswith(('analysis_', 'results_', 'data_', 'inventory_', 'selected_', 'price_data_'))]
        for key in keys_to_clear:
            del st.session_state[key]
        st.session_state.analysis_completed = False
        st.rerun()
    
    # 开始分析按钮
    start_analysis = st.sidebar.button("🚀 开始分析", type="primary")
    
    if start_analysis or st.session_state.analysis_completed:
        if not selected_symbols:
            st.error("请至少选择一个品种进行分析！")
            return
        
        # 如果是新的分析请求，执行分析
        if start_analysis and not st.session_state.analysis_completed:
            st.info(f"正在分析 {len(selected_symbols)} 个品种...")
            
            # 获取数据
            data_dict = get_futures_inventory_data(selected_symbols)
            
            if not data_dict:
                st.error("未获取到任何有效数据，请检查网络连接或稍后重试。")
                return
            
            st.success(f"成功获取 {len(data_dict)} 个品种的数据")
            
            # 分析数据
            analyzer = FuturesInventoryAnalyzer(confidence_level)
            results = []
            inventory_trends = {'累库品种': [], '去库品种': [], '库存稳定品种': []}
            
            for symbol, df in data_dict.items():
                try:
                    category = get_futures_category(symbol)
                    analysis = analyzer.analyze_inventory_trend(df, category)
                    
                    results.append({
                        '品种': symbol,
                        '分类': category,
                        **analysis
                    })
                    
                    trend = analysis['趋势']
                    if trend == '累库':
                        inventory_trends['累库品种'].append(symbol)
                    elif trend == '去库':
                        inventory_trends['去库品种'].append(symbol)
                    else:
                        inventory_trends['库存稳定品种'].append(symbol)
                        
                except Exception as e:
                    st.warning(f"分析 {symbol} 时出错: {str(e)}")
            
            if not results:
                st.error("分析失败，未生成任何结果。")
                return
            
            results_df = pd.DataFrame(results)
            results_df = results_df.sort_values('信号强度', ascending=False)
            
            # 保存到session state
            st.session_state.results_df = results_df
            st.session_state.data_dict = data_dict
            st.session_state.inventory_trends = inventory_trends
            st.session_state.analysis_completed = True
            st.session_state.analysis_mode = analysis_mode
        
        # 从session state获取数据
        results_df = st.session_state.results_df
        data_dict = st.session_state.data_dict
        inventory_trends = st.session_state.inventory_trends
        
        # 检查数据有效性
        if results_df is None or data_dict is None or inventory_trends is None:
            st.error("分析数据丢失，请重新分析。")
            return
        
        # 显示结果
        st.header("📈 分析结果")
        
        # 汇总统计
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("总品种数", len(results_df))
        with col2:
            st.metric("累库品种", len(inventory_trends['累库品种']), 
                     delta=f"{len(inventory_trends['累库品种'])/len(results_df)*100:.1f}%")
        with col3:
            st.metric("去库品种", len(inventory_trends['去库品种']),
                     delta=f"{len(inventory_trends['去库品种'])/len(results_df)*100:.1f}%")
        with col4:
            st.metric("稳定品种", len(inventory_trends['库存稳定品种']),
                     delta=f"{len(inventory_trends['库存稳定品种'])/len(results_df)*100:.1f}%")
        
        # 汇总图表
        st.subheader("📊 汇总图表")
        create_summary_charts(results_df, inventory_trends)
        
        # 详细结果表格
        st.subheader("📋 详细分析结果")
        
        # 筛选选项
        trend_filter = st.selectbox(
            "筛选趋势类型",
            ["全部", "累库", "去库", "稳定"],
            key="trend_filter_selectbox"
        )
        
        if trend_filter != "全部":
            filtered_df = results_df[results_df['趋势'] == trend_filter]
        else:
            filtered_df = results_df
        
        # 格式化显示
        display_df = filtered_df.copy()
        display_df['变化率'] = display_df['变化率'].apply(lambda x: f"{x:.2f}%")
        display_df['信号强度'] = display_df['信号强度'].apply(lambda x: f"{x:.3f}")
        display_df['趋势强度'] = display_df['趋势强度'].apply(lambda x: f"{x:.3f}")
        
        st.dataframe(
            display_df,
            use_container_width=True,
            height=400
        )
        
        # 重点品种分析
        if analysis_mode == "单品种详细分析" and selected_symbols[0] in data_dict:
            st.subheader(f"🔍 {selected_symbols[0]} 详细分析")
            
            symbol = selected_symbols[0]
            df = data_dict[symbol]
            analysis_result = results_df[results_df['品种'] == symbol].iloc[0].to_dict()
            
            # 选择图表类型
            chart_type = st.radio(
                "选择图表类型",
                ["库存走势图", "库存价格对比图"],
                horizontal=True,
                key="chart_type_radio"
            )
            
            if chart_type == "库存走势图":
                # 创建库存趋势图
                fig = create_plotly_trend_chart(df, symbol, analysis_result)
                st.plotly_chart(fig, use_container_width=True, key=f"single_trend_chart_{symbol}_{hash(str(analysis_mode)) % 1000}")
            else:
                # 创建库存价格对比图
                with st.spinner(f"正在获取{symbol}的价格数据..."):
                    price_df = get_futures_price_data(symbol)
                
                if price_df is not None:
                    # 对齐数据时间范围
                    aligned_inventory, aligned_price = align_inventory_and_price_data(df, price_df)
                    
                    # 创建库存价格对比图
                    fig = create_plotly_inventory_price_chart(aligned_inventory, aligned_price, symbol, analysis_result)
                    st.plotly_chart(fig, use_container_width=True, key=f"single_price_chart_{symbol}_{hash(str(analysis_mode)) % 1000}")
                    
                    # 显示数据对齐信息
                    st.info(f"数据时间范围: {aligned_inventory['日期'].min().date()} 到 {aligned_inventory['日期'].max().date()}")
                else:
                    st.warning(f"无法获取{symbol}的价格数据，显示库存走势图")
                    fig = create_plotly_trend_chart(df, symbol, analysis_result)
                    st.plotly_chart(fig, use_container_width=True, key=f"fallback_trend_chart_{symbol}_{hash(str(analysis_mode)) % 1000}")
            
            # 关键指标
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("库存变化率", f"{analysis_result['变化率']:.2f}%")
                st.metric("趋势强度", f"{analysis_result['趋势强度']:.3f}")
            with col2:
                st.metric("信号强度", f"{analysis_result['信号强度']:.3f}")
                st.metric("库存周转率", f"{analysis_result['库存周转率']:.3f}")
            with col3:
                st.metric("平均日变化", f"{analysis_result['平均日变化']:.2f}")
                st.metric("动态阈值", f"{analysis_result['动态阈值']:.2f}")
        
        # 重点关注品种
        if inventory_trends['累库品种'] or inventory_trends['去库品种']:
            st.subheader("⚠️ 重点关注品种")
            
            col1, col2 = st.columns(2)
            
            with col1:
                if inventory_trends['累库品种']:
                    st.markdown("**🟢 累库品种 TOP5**")
                    top_accumulation = results_df[results_df['趋势'] == '累库'].head(5)
                    for _, row in top_accumulation.iterrows():
                        st.write(f"• {row['品种']}: {row['变化率']:.1f}% (信号强度: {row['信号强度']:.2f})")
            
            with col2:
                if inventory_trends['去库品种']:
                    st.markdown("**🔴 去库品种 TOP5**")
                    top_depletion = results_df[results_df['趋势'] == '去库'].head(5)
                    for _, row in top_depletion.iterrows():
                        st.write(f"• {row['品种']}: {row['变化率']:.1f}% (信号强度: {row['信号强度']:.2f})")
            
            # 信号品种库存价格对比分析
            st.subheader("📊 信号品种库存价格对比分析")
            
            # 选择要分析的信号品种
            signal_symbols = inventory_trends['累库品种'] + inventory_trends['去库品种']
            if signal_symbols:
                # 使用session state来保存选择状态
                if 'selected_signal_symbols' not in st.session_state:
                    st.session_state.selected_signal_symbols = signal_symbols[:3] if len(signal_symbols) >= 3 else signal_symbols
                
                selected_signal_symbols = st.multiselect(
                    "选择要查看库存价格对比的品种",
                    signal_symbols,
                    default=st.session_state.selected_signal_symbols,
                    key="signal_symbols_multiselect"
                )
                
                # 更新session state
                st.session_state.selected_signal_symbols = selected_signal_symbols
                
                if selected_signal_symbols:
                    # 使用session state缓存价格数据
                    price_cache_key = f"price_data_{hash(tuple(sorted(selected_signal_symbols)))}"
                    if price_cache_key not in st.session_state:
                        # 批量获取价格数据
                        with st.spinner(f"正在批量获取{len(selected_signal_symbols)}个品种的价格数据..."):
                            st.session_state[price_cache_key] = get_multiple_price_data_streamlit(tuple(selected_signal_symbols))
                    
                    price_data_dict = st.session_state[price_cache_key]
                    
                    for idx, symbol in enumerate(selected_signal_symbols):
                        if symbol in data_dict:
                            with st.expander(f"📈 {symbol} 库存价格对比分析", expanded=False):
                                df = data_dict[symbol]
                                analysis_result = results_df[results_df['品种'] == symbol].iloc[0].to_dict()
                                
                                # 使用预获取的价格数据
                                price_df = price_data_dict.get(symbol)
                                
                                if price_df is not None:
                                    # 对齐数据时间范围
                                    aligned_inventory, aligned_price = align_inventory_and_price_data(df, price_df)
                                    
                                    # 创建库存价格对比图
                                    fig = create_plotly_inventory_price_chart(aligned_inventory, aligned_price, symbol, analysis_result)
                                    # 使用索引和品种名确保唯一性
                                    unique_key = f"signal_price_chart_{idx}_{symbol}_{hash(tuple(selected_signal_symbols)) % 10000}"
                                    st.plotly_chart(fig, use_container_width=True, key=unique_key)
                                    
                                    # 显示关键指标
                                    col1, col2, col3, col4 = st.columns(4)
                                    with col1:
                                        st.metric("库存变化率", f"{analysis_result['变化率']:.1f}%")
                                    with col2:
                                        st.metric("信号强度", f"{analysis_result['信号强度']:.2f}")
                                    with col3:
                                        st.metric("趋势强度", f"{analysis_result['趋势强度']:.2f}")
                                    with col4:
                                        # 计算价格变化率
                                        if len(aligned_price) > 1:
                                            price_change_rate = (aligned_price['价格'].iloc[-1] - aligned_price['价格'].iloc[0]) / aligned_price['价格'].iloc[0] * 100
                                            st.metric("价格变化率", f"{price_change_rate:.1f}%")
                                        else:
                                            st.metric("价格变化率", "N/A")
                                else:
                                    st.warning(f"无法获取{symbol}的价格数据")
        
        # 下载结果
        st.subheader("💾 下载结果")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            # CSV下载
            csv = results_df.to_csv(index=False, encoding='utf-8-sig')
            st.download_button(
                label="📄 下载分析结果 (CSV)",
                data=csv,
                file_name=f"期货库存分析结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                mime="text/csv"
            )
        
        with col2:
            # Excel下载
            try:
                with st.spinner("正在生成Excel报告..."):
                    wb = create_excel_report(results_df, inventory_trends, data_dict)
                    excel_buffer = io.BytesIO()
                    wb.save(excel_buffer)
                    excel_buffer.seek(0)
                
                st.download_button(
                    label="📊 下载Excel报告",
                    data=excel_buffer.getvalue(),
                    file_name=f"期货库存分析报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"生成Excel报告失败: {str(e)}")
                st.info("请确保已安装 openpyxl 库: pip install openpyxl")
        
        with col3:
            # 图表下载
            try:
                with st.spinner("正在生成图表包..."):
                    charts_zip = create_charts_zip(results_df, inventory_trends, data_dict)
                
                st.download_button(
                    label="📈 下载图表包 (ZIP)",
                    data=charts_zip,
                    file_name=f"期货库存分析图表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                    mime="application/zip"
                )
            except Exception as e:
                st.error(f"生成图表包失败: {str(e)}")
                st.info("请确保已安装 kaleido 库: pip install kaleido")
        
        # 下载说明
        st.info("""
        📋 **下载说明：**
        - **CSV文件**: 包含所有分析结果的原始数据，适合进一步数据处理
        - **Excel报告**: 包含格式化的分析报告，包括汇总统计、详细结果、重点品种和部分原始数据
        - **图表包**: 包含所有重要图表的PNG文件，包括汇总图表和重点品种趋势图
        """)

if __name__ == "__main__":
    main() 
